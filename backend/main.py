from fastapi import FastAPI, HTTPException, UploadFile, File
from openpyxl import Workbook, load_workbook
from tempfile import NamedTemporaryFile
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware

from typing import List, Union

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

@app.get('/')
def read_root():
    return {"Hello": "World"}

@app.post("/generate_excel")
async def generate_excel(data: List[List[Union[str, int]]]):
    if not data:
        raise HTTPException(status_code=400, detail="No data provided")

    # Create a new Excel workbook
    wb = Workbook()
    ws = wb.active

    # Add data to the workbook
    for row in data:
        ws.append(row)

    # Save the workbook to a temporary file
    with NamedTemporaryFile(delete=False) as tmp:
        file_path = tmp.name
        wb.save(file_path)

        # Return the Excel file as a streaming response
        return StreamingResponse(open(file_path, "rb"), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    
@app.post("/upload_excel")
async def upload_excel(file: UploadFile = File(...)):
    # Save the uploaded Excel file to a temporary location
    with NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp:
        file_content = await file.read()
        tmp.write(file_content)
        tmp.close()

    # Load the Excel file
    wb = load_workbook(tmp.name)
    ws = wb.active

    # Initialize an empty list to store the data
    excel_data = []
    # Iterate over the rows in the worksheet and extract data
    for row in ws.iter_rows(values_only=True):
        excel_data.append(list(row))

    return JSONResponse(content=excel_data)